from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from copy import copy
import copy
import openpyxl
import os


def read_excel_data(file_path):
    result = {}
    # 以只读模式加载工作簿（提高性能）
    wb = openpyxl.load_workbook(file_path, read_only=True)

    # 获取前两个Sheet（索引0和1）
    sheets = wb.worksheets[:2]

    for sheet in sheets:
        # 遍历从第13行开始的行，C列到S列（min_col=3对应C列，max_col=19对应S列）
        for row in sheet.iter_rows(
                min_row=13,
                min_col=3,
                max_col=19,
                values_only=True  # 直接获取单元格值，而非单元格对象
        ):
            key = row[0]  # C列的值（row[0]对应第一个单元格）
            if key is None:
                continue  # 跳过空键

            # 提取D到V列的值（从索引1到19）
            values = list(row[1:])
            result[key] = values

    wb.close()  # 显式关闭工作簿（尤其是read_only模式）
    return result


def find_start_row(sheet):
    """找到第一列首次出现值为1的行号（从1开始）"""
    for row in sheet.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        # 处理合并单元格：实际值在合并区域的左上角单元格
        if cell.value == 1:
            return cell.row
        # 检查单元格是否在合并区域中且合并区域的第一个单元格值为1
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_col == 1 and merged_range.max_col == 1:
                if cell.row in range(merged_range.min_row, merged_range.max_row + 1):
                    first_cell = sheet.cell(row=merged_range.min_row, column=1)
                    if first_cell.value == 1:
                        return merged_range.min_row
    return None  # 如果未找到1，返回None（可根据需求抛出异常）


def read_merged_excel(file_path):
    """读取Excel表，从第一列第一个1的行开始"""
    wb = load_workbook(filename=file_path, data_only=True)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 找到起始行
        start_row = find_start_row(sheet)
        if start_row is None:
            print(f"警告: Sheet '{sheet_name}' 的第一列未找到值1，跳过处理")
            continue

        # 获取合并区域信息（仅第一列）
        merged_ranges = [r for r in sheet.merged_cells.ranges if r.min_col == 1 and r.max_col == 1]

        # 构建行号到合并值的映射
        merge_map = {}
        for merged_range in merged_ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            value = sheet.cell(min_row, 1).value
            for row in range(min_row, max_row + 1):
                merge_map[row] = value

        # 读取数据：从start_row开始
        data_rows = []
        compare_data_rows=[]
        for row_idx in range(start_row, sheet.max_row + 1):
            # 处理第一列的值（优先取合并映射中的值）
            first_col_val = merge_map.get(row_idx, sheet.cell(row_idx, 1).value)
            row_data = [first_col_val]
            # 读取其他列
            for col_idx in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            if row_data[0] != None:
                data_rows.append(row_data)
        # 添加测试过的行
        current_num = 1
        production_code = None
        list_to_insert = None
        i = 0
        while (i < len(data_rows)):
            if data_rows[i][0] == current_num and production_code == None:
                production_code = data_rows[i][3]
                list_to_insert = copy.deepcopy(after_data_dict.get(production_code))
            elif i == len(data_rows) - 1 or data_rows[i + 1][0] != current_num:
                if list_to_insert != None:
                    list_to_insert.insert(0, current_num)
                    list_to_insert.insert(1, None)
                    list_to_insert.insert(2, None)
                    list_to_insert.insert(3, '实测指标')
                    list_to_insert.insert(4, None)
                    list_to_insert.append(None)
                    compare_data_rows.append(data_rows[i])
                    compare_data_rows.append(list_to_insert)
                    list_to_insert = None
                    production_code = None
                current_num = current_num + 1
            i = i + 1

        sheets_data[sheet_name] = {
            'data': compare_data_rows,
            'original_sheet': sheet  # 用于复制列宽
        }

    return sheets_data


def write_merged_excel(sheets_data, output_path):

    for sheet_name, data in sheets_data.items():
        """生成新Excel表，保留第一列合并格式"""
        if len(data['data'])==0:
            continue
        wb_new = Workbook()
        default_sheet = wb_new.active
        wb_new.remove(default_sheet)
        original_sheet = data['original_sheet']
        ws_new = wb_new.create_sheet(sheet_name)
        data_rows = data['data']
        # 写入数据到新表（从新表的第1行开始）
        for row_idx, row_data in enumerate(data_rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws_new.cell(row=row_idx, column=col_idx, value=value)
        # 处理第一列合并：连续相同值合并
        current_value = None
        start_merge_row = 1
        for row_idx, row_data in enumerate(data_rows, start=1):
            value = row_data[0]
            if current_value is None:
                current_value = value
                start_merge_row = row_idx
            else:
                if value != current_value:
                    if start_merge_row < row_idx - 1:
                        ws_new.merge_cells(
                            start_row=start_merge_row, end_row=row_idx - 1,
                            start_column=1, end_column=1
                        )
                    current_value = value
                    start_merge_row = row_idx
        # 处理最后一段合并
        if current_value is not None and start_merge_row < len(data_rows):
            ws_new.merge_cells(
                start_row=start_merge_row, end_row=len(data_rows),
                start_column=1, end_column=1
            )
        # 复制原表的列宽
        for col in original_sheet.columns:
            col_letter = get_column_letter(col[0].column)
            ws_new.column_dimensions[col_letter].width = original_sheet.column_dimensions[col_letter].width
            wb_new.save(output_path+sheet_name.replace('.','_')+'.xlsx')
        print(f"处理完成，新文件已保存至：{output_path + sheet_name.replace('.', '_') + '.xlsx'}")


# 使用示例
if __name__ == "__main__":

    before_folder_path = '../5月份数据/before'  # 处理前的预测数据
    out_folder_path = '../compare_out_files_by_sheet_5'  # 结果保存文件路径
    after_file_path = "../5月份数据/after/2025年05月海绵钛二期GTB.xlsx"  # 处理后的实测数据
    after_data_dict = read_excel_data(after_file_path)
    # 遍历文件夹中的所有文件和子文件夹
    for filename in os.listdir(before_folder_path):
        file_path = os.path.join(before_folder_path, filename)
        if os.path.isfile(file_path):  # 只处理文件，忽略子文件夹
            # 读取数据并处理
            sheets_data = read_merged_excel(file_path)
            # 写入新表
            output_file = os.path.join(out_folder_path, 'output_')
            write_merged_excel(sheets_data, output_file)

