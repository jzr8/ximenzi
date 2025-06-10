import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from stove import Stove
from draw import *
from method.EWMA import *
from method.window_LR import *
from method.kalman_filter import *
from method.kalman_EWMA import *


# 读取一个日期的excel表格 将其中的每个炉次提取为Stove对象  返回Stove对象的列表
def read_singel_excel(path):
    df = pd.read_excel(path, header=None)  # 无标题
    filename = os.path.basename(path)  # 获取文件名
    name_without_ext = os.path.splitext(filename)[0]  # 去掉扩展名
    date = name_without_ext.split('_')[-2] + '_' + name_without_ext.split('_')[-1]  # 提取日期
    ST = []  # 保存stove对象

    num = len(df) // 2  # 该表格下的总序号数
    # 遍历每个序列
    for i in range(num):
        pre_row = 2 * i  # 预测指标行
        real_row = pre_row + 1  # 实测指标行
        pre_dict = {}  # 存放预测指标
        real_dict = {}  # 存放实测指标

        # 预测指标
        pre_dict['grade'] = df.iloc[pre_row, 6]  # 等级
        pre_dict['Fe'] = df.iloc[pre_row, 9]  # Fe
        pre_dict['Cl'] = df.iloc[pre_row, 11]  # Cl
        pre_dict['C'] = df.iloc[pre_row, 12]  # C
        pre_dict['N'] = df.iloc[pre_row, 13]  # N
        pre_dict['O'] = df.iloc[pre_row, 14]  # O
        pre_dict['Ni'] = df.iloc[pre_row, 18]  # Ni
        pre_dict['Cr'] = df.iloc[pre_row, 19]  # Cr
        pre_dict['hard'] = df.iloc[pre_row, 20]  # 布什硬度

        # 实测指标
        real_dict['grade'] = df.iloc[real_row, 6]  # 等级
        real_dict['Fe'] = df.iloc[real_row, 9]  # Fe
        real_dict['Cl'] = df.iloc[real_row, 11]  # Cl
        real_dict['C'] = df.iloc[real_row, 12]  # C
        real_dict['N'] = df.iloc[real_row, 13]  # N
        real_dict['O'] = df.iloc[real_row, 14]  # O
        real_dict['Ni'] = df.iloc[real_row, 18]  # Ni
        real_dict['Cr'] = df.iloc[real_row, 19]  # Cr
        real_dict['hard'] = df.iloc[real_row, 20]  # 布什硬度

        st = Stove(date, i + 1, pre_dict, real_dict)  # 构造Stove对象
        ST.append(st)

    return ST


# 获取所有指定的数据  将不同日期不同序号的数据 封装在Stove对象中 再保存在ST列表中
def get_ST(input_get, month_get, date_get):
    st_get = []
    for file in input_get.rglob('*.xlsx'):  # 只寻找xlsx文件
        filename = os.path.basename(file)  # 获取文件名
        name_without_ext = os.path.splitext(filename)[0]  # 去掉扩展名
        m = name_without_ext.split('_')[-2]  # 提取月份
        d = name_without_ext.split('_')[-1]  # 提取日期
        if m not in month_get:
            continue
        if d not in date_get:
            continue

        file = './' + str(file)
        # 读取该表格中的数据
        st_get.extend(read_singel_excel(file))
    return st_get


# 对炉基于等级进行筛选，只保留指定等级的炉子
def grade_filt(get_ST, get_grade):
    new_ST = []
    for st in get_ST:
        if st.grade[1] in get_grade:
            new_ST.append(st)
    return new_ST


# 对输入的Stove列表 进行数据统计
def stat(st_list, get_stand):
    # 符合相应指标的炉子数量
    stat_accord = {'Fe': 0, 'Cl': 0, 'C': 0, 'N': 0, 'O': 0, 'Ni': 0, 'Cr': 0, 'hard': 0, 'stand': 0, 'grade': 0}
    # 相应指标的偏差总和
    stat_error_sum = {'Fe': 0, 'Cl': 0, 'C': 0, 'N': 0, 'O': 0, 'Ni': 0, 'Cr': 0, 'hard': 0}
    # 相应指标的平均偏差
    stat_error_ave = {'Fe': 0, 'Cl': 0, 'C': 0, 'N': 0, 'O': 0, 'Ni': 0, 'Cr': 0, 'hard': 0}
    # 相应指标的偏差最大值  这里找到最大的10炉数据 按照升序排序 0索引代表最小的数据  9索引代表最大的数据
    stat_error_max = {'Fe': [0] * 10, 'Cl': [0] * 10, 'C': [0] * 10, 'N': [0] * 10, 'O': [0] * 10, 'Ni': [0] * 10,
                      'Cr': [0] * 10, 'hard': [0] * 10}
    stat_Stove_item_max = {'Fe': [None] * 10, 'Cl': [None] * 10, 'C': [None] * 10, 'N': [None] * 10, 'O': [None] * 10,
                           'Ni': [None] * 10, 'Cr': [None] * 10, 'hard': [None] * 10}

    st_num = len(st_list)  # 得到总炉数
    # 遍历每个炉数据
    for st in st_list:
        st.judge_error(get_stand)  # 根据设定的标准偏差进行偏差判断

        # 符合标准的数量 & 符合等级的数量
        if -1 not in [st.Fe[0], st.Cl[0], st.C[0], st.N[0], st.O[0], st.Ni[0], st.Cr[0], st.hard[0]]:
            stat_accord['stand'] += 1
        if st.grade[0] == st.grade[1]:
            stat_accord['grade'] += 1

        # 各个指标符合的数量
        for key in stat_accord:
            if key != 'stand' and key != 'grade':
                stat_accord[key] += 1 if getattr(st, key)[0] == 1 else 0

        # 各个指标的偏差总和
        for key in stat_error_sum:
            stat_error_sum[key] += getattr(st, key)[3]

        # 各个指标的偏差最大值  对该炉的数据 进行判断排序
        for key in stat_error_max:
            if getattr(st, key)[3] >= stat_error_max[key][0]:
                # 添加到列表中
                stat_error_max[key].append(getattr(st, key)[3])
                stat_Stove_item_max[key].append(st)
                combined = list(zip(stat_error_max[key], stat_Stove_item_max[key]))  # 将两个列表打包为一对一元组列表
                combined.sort(key=lambda x: x[0])  # 按数值升序排序
                combined.pop(0)  # 删除最小值（排在第一位的元素）
                new_nums, new_labels = zip(*combined)  # 解包为两个新列表
                # 转换为列表类型（因为 zip 返回的是元组）
                stat_error_max[key] = list(new_nums)
                stat_Stove_item_max[key] = list(new_labels)

                # stat_error_max[key] = getattr(st, key)[3]
                # stat_Stove_item_max[key] = st

    # # 找到最大值的生产日期和批次
    # for key in stat_Stove_item_max:
    #     print(key+f':date:{stat_Stove_item_max[key].date}--num:{stat_Stove_item_max[key].number}')

    # 找到偏差最大的那几个炉子的等级
    # for key in stat_Stove_item_max:
    #     print(key + f':{[stat_Stove_item_max[key][i].grade[1] for i in range(len(stat_Stove_item_max[key]))]}')

    # 计算各指标偏差的平均值
    for key in stat_error_ave:
        stat_error_ave[key] = stat_error_sum[key] / st_num

    # 符合率
    keys_to_extract = ['Fe', 'Cl', 'C', 'N', 'O', 'Ni', 'Cr', 'hard']
    stat_accord_p = {k: stat_accord[k] / st_num for k in keys_to_extract if k in stat_accord}
    # 偏差最大值
    data_max = [stat_error_max[key][-1] for key in stat_error_max]
    data_max = [round(num, 3) for num in data_max]  # 保留小数点后三位
    # 平均偏差
    data_error_ave = [stat_error_ave[key] for key in stat_error_ave]
    data_error_ave = [round(num, 3) for num in data_error_ave]  # 保留小数点后三位

    print(f'炉子总数为：{st_num}')

    # 平均偏差  最大偏差  符合数量  符合率  标准符合率  等级符合率
    return data_error_ave, data_max, stat_accord, stat_accord_p, stat_accord['stand'] / st_num, stat_accord[
        'grade'] / st_num


# 设置各个元素想要达到的指标，逐步增加标准偏差，使得达到指定值
def set_accord():
    # 设置符合率  没要求就设置为0即可
    stand_p = {'Fe': 0, 'Cl': 0, 'C': 0, 'N': 0, 'O': 0, 'Ni': 0, 'Cr': 0, 'hard': 0}

    # 更改标准偏差 以达到指定的符合率（如果设置了的话）  检测符合率是否到达指定值
    while any([acc_p[key] < stand_p[key] for key in stand_p]):
        for key in stand:
            if acc_p[key] < stand_p[key]:
                stand[key] += 0.001

        # 进行数据统计
        d_error_ave, d_max, acc, acc_p, st_per, gr_per = stat(ST, stand)


# 将结果写入excel表格中
def write_excel(data_error_ave, data_max, accord, accord_p, st_percent, gr_percent, get_stand):
    # 构造一个 DataFrame，并设置行索引
    accord_pp = [accord_p['Fe'], accord_p['Cl'], accord_p['C'], accord_p['N'], accord_p['O'], accord_p['Ni'],
                 accord_p['Cr'], accord_p['hard']]
    data = {
        '偏差标准≤': [get_stand['Fe'], get_stand['Cl'], get_stand['C'], get_stand['N'], get_stand['O'], get_stand['Ni'],
                      get_stand['Cr'], get_stand['hard']],
        '平均偏差': data_error_ave,
        '最大值': data_max,
        '符合炉次': [accord['Fe'], accord['Cl'], accord['C'], accord['N'], accord['O'], accord['Ni'],
                     accord['Cr'], accord['hard'], ],
        '符合率%': [f"{x * 100:.1f}%" for x in accord_pp],
        '判断标准符合率%': [f"{st_percent * 100:.1f}%"],
        '等级符合率%': [f"{gr_percent * 100:.1f}%"],
    }
    df = pd.DataFrame(data, index=['Fe（%）', 'Cl（%）', 'C（%）', 'N（%）', 'O（%）', 'Ni（%）', 'Cr（%）', '布什硬度'])

    # 写入 Excel，设置行索引列标题为 "姓名"
    df.to_excel(output, index=True, index_label='项目')

    # 加载该 Excel 文件进行格式修改
    wb = load_workbook(output)
    ws = wb.active  # 默认是第一个工作表

    # 设置所有单元格的字体大小
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=14)  # 设置字体大小为14

    # 合并单元格
    ws.merge_cells('G2:G9')
    ws.merge_cells('H2:H9')

    # 设置对齐方式（水平+垂直居中）
    alignment = Alignment(horizontal='center', vertical='center')
    ws['G2'].alignment = alignment
    ws['H2'].alignment = alignment

    wb.save(output)


# 对输入的炉列表 进行相关偏差数据的统计 最终输出excel结果
def output_data(get_ST, get_stand):
    # 进行数据统计
    # 得到 平均偏差  最大偏差  符合数量  符合率  标准符合率  等级符合率
    d_error_ave, d_max, acc, acc_p, st_per, gr_per = stat(get_ST, get_stand)
    # 写入excel表格中
    write_excel(d_error_ave, d_max, acc, acc_p, st_per, gr_per, get_stand)


if __name__ == '__main__':
    # 注意：运行该代码时 不要打开相应的输入文件 会报错
    input = Path('../compare_out_files_by_sheet/')  # 输入文件夹路径
    output = '../线掺配预测表2025.4_2.xlsx'  # 输出结果路径
    month = ['4']  # 提取表格的月份
    date = [str(i) for i in range(1, 32)]  # 提取表格的日期 想要提取1~31日的，则设置为(1，32)
    # 得到指定日期的Stove列表
    ST = get_ST(input, month, date)

    # 筛选等级
    # set_grade = ['0级']
    # ST = grade_filt(ST, set_grade)

    # 设置标准偏差
    stand = {'Fe': 0.009, 'Cl': 0.010, 'C': 0.001, 'N': 0.001, 'O': 0.005, 'Ni': 0.004, 'Cr': 0.005, 'hard': 2.0}

    '''------------------------输出excel表格------------------------'''
    # output_data(ST, stand)
    '''-----------------------------------------------------------'''

    '''----------------------相关曲线图绘制-------------------------'''
    # draw_error_by_date(ST, 'Fe')  # 对指定元素绘制偏差曲线  按天
    # draw_error_by_number(ST, 'Cl')  # 按炉次

    '''-----------------------------------------------------------'''

    '''----------------------EWMA(指数加权移动平均)------------------'''
    # correct_opti_one(ST, stand, 'Fe')  # 固定一个参数 找另一个参数最优（画曲线图）

    # correct_opti_two(ST, stand, 'Fe')  # 直接通过遍历找出两个参数的最优组合

    # 设置指定的阈值（gate）、权重（theta），返回指定元素的符合率（通过EWMA修正后）
    # EWMA_correct(ST, stand, 'Fe', gate=0.009, theta=0.25)

    '''-----------------------------------------------------------'''

    '''------------------------WLR(滑动窗口线性回归)------------------'''
    # 设置指定的阈值、滑动窗口数，返回指定元素的符合率（通过WLR修正后）
    # WLR_correct(ST, stand, 'Fe', gate=0.009, window_size=9)

    # 通过遍历得到最优window_size
    # opti_window_size(ST, stand, 'Fe')

    '''----------------------------------------------------------- '''

    '''---------------------------kalman---------------------------'''
    # 通过kalman滤波预测进行修正后 得到的指定元素的符合率
    # kalman_correct(ST, stand, 'Fe', gate=0.009)

    '''------------------------------------------------------------'''

    '''------------------------kalman+EWMA-------------------------'''
    # kalman+EWMA预测的均值 作为 最终预测偏差
    # kalman_EWMA_correct(ST, stand, 'Fe', gate=0.009, theta=0.75)

    # 遍历得到最优参数
    # kalman_EWMA_opti_two(ST, stand, 'Fe')

    '''------------------------------------------------------------'''
